"""updt_input_excel.py
======================
Update optimizer_config.xlsx when the balance sheet product list changes.

Run this whenever you re-run the extract_params pipeline (new products added,
removed, or balance percentages changed).

Behaviour
---------
- Existing products: current_pct is refreshed; user-set fixed/lb_pct/ub_pct/notes are PRESERVED.
- New products:      added with default bounds (fixed=FALSE, lb=0, ub=100).
- Removed products:  dropped from the sheet.
- T1 capital row:    refreshed from equity product balances (if set to 0 / auto).

The config sheet (global params) is never touched — user settings are preserved.

Usage:
    python updt_input_excel.py
"""
import os, sys

_HERE    = os.path.dirname(os.path.abspath(__file__))
_OPTPREP = os.path.normpath(os.path.join(_HERE, "..", "..", "optimize_prep", "python_code"))
for _p in (_HERE, _OPTPREP):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from bs_vector import BalanceSheetParams
from optimizer_io import create_optimizer_excel, _CONFIG_ROWS

_PARAMS = os.path.normpath(os.path.join(_HERE, "..", "..", "optimize_prep", "output", "product_params.npz"))
_XL     = os.path.normpath(os.path.join(_HERE, "..", "input", "optimizer_config.xlsx"))


def update_optimizer_excel(
    params_path: str = _PARAMS,
    excel_path:  str = _XL,
) -> None:
    """Refresh optimizer_config.xlsx while preserving user-configured bounds.

    Parameters
    ----------
    params_path : path to product_params.npz
    excel_path  : path to optimizer_config.xlsx
    """
    params = BalanceSheetParams.load(params_path)
    ta     = params.total_assets

    # ── Build current product list from params ────────────────────────────────
    current: dict[tuple[str, str], dict] = {}
    for pc, side, name, ccy, bal in zip(
        params.product_code, params.bs_side, params.product_name,
        params.currency, params.balance_arr,
    ):
        key = (str(pc), str(side))
        if key not in current:
            current[key] = {
                "product_code": str(pc),
                "bs_side":      str(side),
                "product_name": str(name),
                "currency":     str(ccy),
                "current_pct":  0.0,
            }
        current[key]["current_pct"] += float(bal) / ta * 100.0

    t1_from_params = float(params.balance_arr[params.is_equity].sum())

    # ── If Excel doesn't exist yet, create it from scratch ───────────────────
    if not os.path.exists(excel_path):
        print(f"Excel not found — creating from scratch: {excel_path}")
        create_optimizer_excel(params, excel_path)
        return

    # ── Read existing user-set bounds ─────────────────────────────────────────
    try:
        existing_df = pd.read_excel(excel_path, sheet_name="products", header=1)
    except Exception as e:
        print(f"Could not read products sheet: {e}")
        print("Recreating Excel from scratch.")
        create_optimizer_excel(params, excel_path)
        return

    existing_df.columns = [str(c).strip().lower().replace(" ", "_") for c in existing_df.columns]
    _remap = {
        "product_code": "product_code",
        "side":         "bs_side",
        "fixed":        "fixed",
        "lb_%":         "lb_pct",
        "ub_%":         "ub_pct",
        "notes":        "notes",
    }
    existing_df.rename(columns=_remap, inplace=True)
    existing_df["product_code"] = existing_df["product_code"].astype(str).str.strip()
    existing_df["bs_side"]      = existing_df["bs_side"].astype(str).str.strip()
    existing_df = existing_df[existing_df["bs_side"].isin(["A", "L", "E"])].copy()

    # Build lookup: (pc, side) -> {fixed, lb_pct, ub_pct, notes}
    user_settings: dict[tuple[str, str], dict] = {}
    for _, row in existing_df.iterrows():
        key = (str(row["product_code"]), str(row["bs_side"]))
        fixed_raw = row.get("fixed", False)
        if isinstance(fixed_raw, str):
            is_fixed = fixed_raw.strip().upper() == "TRUE"
        else:
            is_fixed = bool(fixed_raw)
        try:
            lb = float(row.get("lb_pct", 0.0))
            ub = float(row.get("ub_pct", 100.0))
        except (TypeError, ValueError):
            lb, ub = 0.0, 100.0
        user_settings[key] = {
            "fixed":  is_fixed,
            "lb_pct": lb,
            "ub_pct": ub,
            "notes":  str(row.get("notes", "")) if pd.notna(row.get("notes", "")) else "",
        }

    # ── Merge: refresh current_pct, preserve user settings ───────────────────
    new_rows = []
    for key, prod in current.items():
        pc, side = key
        cpct = round(prod["current_pct"], 4)
        is_equity = side == "E"

        if key in user_settings:
            u = user_settings[key]
            row = {
                "product_code": pc,
                "bs_side":      side,
                "product_name": prod["product_name"],
                "currency":     prod["currency"],
                "current_pct":  cpct,
                "fixed":        u["fixed"],
                "lb_pct":       u["lb_pct"],
                "ub_pct":       u["ub_pct"],
                "notes":        u["notes"],
            }
        else:
            # New product — default bounds
            row = {
                "product_code": pc,
                "bs_side":      side,
                "product_name": prod["product_name"],
                "currency":     prod["currency"],
                "current_pct":  cpct,
                "fixed":        True if is_equity else False,
                "lb_pct":       cpct if is_equity else 0.0,
                "ub_pct":       cpct if is_equity else 100.0,
                "notes":        "equity — always pinned" if is_equity else "[NEW PRODUCT]",
            }

        new_rows.append(row)

    # Report changes
    old_keys  = set(user_settings.keys())
    new_keys  = set(current.keys())
    added     = new_keys - old_keys
    removed   = old_keys - new_keys
    if added:
        print(f"  + Added products:   {sorted(added)}")
    if removed:
        print(f"  - Removed products: {sorted(removed)}")
    if not added and not removed:
        print("  No product list changes — refreshing current_pct values only.")

    # ── Rewrite the products sheet in place ──────────────────────────────────
    wb = openpyxl.load_workbook(excel_path)

    # ── Patch config sheet: update T1 + append any missing rows ─────────────
    ws_cfg = wb["config"]

    # Collect existing key → (row_number, value_cell)
    existing_keys: dict[str, int] = {}
    for row_cells in ws_cfg.iter_rows():
        for cell in row_cells:
            if cell.column == 4 and cell.value is not None:
                existing_keys[str(cell.value)] = cell.row

    # Update T1 if it was left at 0 (auto)
    t1_row = existing_keys.get("tier1_capital")
    if t1_row is not None:
        val_cell = ws_cfg.cell(row=t1_row, column=2)
        if float(val_cell.value or 0) == 0.0:
            val_cell.value = t1_from_params
            print(f"  T1 capital (auto): {t1_from_params/1e6:,.0f}M PLN")

    # Append missing config rows (new feature params added to _CONFIG_ROWS)
    _GREY_HDR  = "D6DCE4"
    _LT_BLUE   = "DEEAF1"
    _WHITE     = "FFFFFF"
    _THIN      = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"),  bottom=Side(style="thin"))
    next_row = ws_cfg.max_row + 1
    added_keys = []
    for key, label, default_val, desc in _CONFIG_ROWS:
        if key not in existing_keys:
            ws_cfg.cell(row=next_row, column=1, value=label).border = _THIN
            ws_cfg.cell(row=next_row, column=1).fill = PatternFill("solid", fgColor=_GREY_HDR)
            val_cell = ws_cfg.cell(row=next_row, column=2, value=default_val)
            val_cell.fill   = PatternFill("solid", fgColor=_LT_BLUE)
            val_cell.border = _THIN
            val_cell.alignment = Alignment(horizontal="left")
            ws_cfg.cell(row=next_row, column=3, value=desc).border = _THIN
            ws_cfg.cell(row=next_row, column=3).fill = PatternFill("solid", fgColor=_WHITE)
            ws_cfg.cell(row=next_row, column=4, value=key)
            ws_cfg.column_dimensions["D"].hidden = True
            added_keys.append(key)
            next_row += 1
    if added_keys:
        print(f"  Config sheet: added missing rows: {added_keys}")
    else:
        print("  Config sheet: no missing rows to add.")

    # Replace products sheet (keep it at index 1, after config)
    del wb["products"]
    ws_prod = wb.create_sheet("products", 1)

    # Re-import write helper (same module)
    from optimizer_io import _write_products_sheet
    _write_products_sheet(ws_prod, new_rows)

    wb.save(excel_path)
    print(f"Updated: {excel_path}  ({len(new_rows)} products)")


if __name__ == "__main__":
    print("Updating optimizer_config.xlsx...")
    update_optimizer_excel()
