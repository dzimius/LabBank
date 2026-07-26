"""optimizer_io.py
==================
Excel I/O for balance sheet optimizer configuration.

Workflow
--------
1. create_optimizer_excel(params, path) — write template with current values
2. User edits Excel: set mode, global params, per-product bounds
3. load_optimizer_config(path, params)  — parse → OptimizationConfig

Sheets
------
config      : global parameters (mode, T1, SOT floors, shift %, etc.)
products    : per-product bounds (lb_pct, ub_pct, fixed flag)

Mode semantics (same as OptimizationConfig)
--------------------------------------------
full        : all non-equity free in [0,100%]; fixed=TRUE pins a product
partial     : only products with fixed=FALSE are optimised; rest pinned
max_shift   : each free product bounded by [current ± max_shift_pct]; fixed=TRUE pins
custom      : lb_pct / ub_pct columns used directly as asymmetric bounds per product
"""
from __future__ import annotations

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# bs_vector lives in optimize_prep/python_code; bs_optimizer is in the same directory
_HERE    = os.path.dirname(os.path.abspath(__file__))
_OPTPREP = os.path.normpath(os.path.join(_HERE, "..", "..", "optimize_prep", "python_code"))
for _p in (_HERE, _OPTPREP):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from bs_vector import BalanceSheetParams
from bs_optimizer import OptimizationConfig


# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────
_BLUE_HDR  = "1F4E79"    # dark navy header
_GREY_HDR  = "D6DCE4"    # light grey sub-header
_ORANGE    = "F4B942"    # highlight equity rows
_GREEN_HDR = "375623"    # products header
_LT_GREEN  = "E2EFDA"    # editable rows background
_LT_BLUE   = "DEEAF1"    # config value background
_WHITE     = "FFFFFF"

_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _hdr_font(white: bool = True) -> Font:
    return Font(bold=True, color=_WHITE if white else "000000", size=10)


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


def _set_col_width(ws, col_letter: str, width: float) -> None:
    ws.column_dimensions[col_letter].width = width


# ─────────────────────────────────────────────────────────────────────────────
# Config sheet
# ─────────────────────────────────────────────────────────────────────────────

_CONFIG_ROWS = [
    # (key, label, default_value, description)
    ("mode",            "Optimization mode",           "max_shift",
     "full | partial | max_shift | custom"),
    ("tier1_capital",   "Tier 1 capital (PLN)",         0.0,
     "T1 capital — denominator for EVE and NII SOT %.  0 = auto-derive from equity product balances."),
    ("sot_eve_floor",   "EVE SOT floor (% T1)",         -15.0,
     "EBA regulatory lower bound on delta_EVE / T1.  E.g. set to -17 with buffer 0 instead of -15 with buffer 2."),
    ("sot_nii_floor",   "NII SOT floor (% T1)",         -5.0,
     "EBA regulatory lower bound on delta_NII / T1."),
    ("sot_eve_buffer",  "EVE SOT buffer (% T1)",         3.0,
     "Relax EVE floor by this amount to absorb method B approx error (~2.7% T1).  Effective floor = floor - buffer."),
    ("sot_nii_buffer",  "NII SOT buffer (% T1)",         0.0,
     "Relax NII floor by this amount to absorb method B approx error (~0.7% T1)."),
    ("include_irs",     "Include IR derivatives (IRS)",  True,
     "TRUE = count IRS EVE/NII in constraints (hedged view, default).  "
     "FALSE = exclude IRS from all metric computations (unhedged banking book only).  "
     "IRS notional is always pinned regardless of this setting."),
    ("min_lcr",         "Minimum LCR",                  1.0,
     "Minimum acceptable LCR ratio per currency"),
    ("min_nsfr",        "Minimum NSFR",                 1.0,
     "Minimum acceptable NSFR ratio per currency"),
    ("min_t1_rwa",      "Min T1/RWA ratio",             0.18,
     "CET1 floor: T1 / RWA >= this value (e.g. 0.18 = 18%).  0 = no constraint.  "
     "Baseline T1/RWA ≈ 44% (T1=1,600M, RWA=3,620M).  "
     "LP optimal (max loans) reaches ~18% — set this to 0.18 to make constraint binding."),
    ("max_shift_pct",   "Max shift per product (%TA)",   3.0,
     "Used in max_shift mode: max weight change per product as % of total_assets"),
    ("max_iter",        "Max optimizer iterations",      500,
     "SLSQP maximum iterations (increase if 'Iteration limit exceeded')"),
]


def _write_config_sheet(ws, defaults: dict) -> None:
    """Write the 'config' sheet with global optimizer parameters."""
    # Title
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "Optimizer Global Configuration"
    title.font  = Font(bold=True, size=13, color=_WHITE)
    title.fill  = _fill(_BLUE_HDR)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Column headers
    headers = ["Parameter Key", "Value", "Description"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font  = _hdr_font()
        cell.fill  = _fill(_GREEN_HDR)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    # Data rows
    for r, (key, label, default_val, desc) in enumerate(_CONFIG_ROWS, start=3):
        ws.cell(row=r, column=1, value=label).border = _THIN_BORDER
        ws.cell(row=r, column=1).fill = _fill(_GREY_HDR)
        ws.cell(row=r, column=1).font = Font(bold=False, size=10)

        val_cell = ws.cell(row=r, column=2, value=defaults.get(key, default_val))
        val_cell.fill   = _fill(_LT_BLUE)
        val_cell.border = _THIN_BORDER
        val_cell.alignment = Alignment(horizontal="left")
        # Store the key as a row comment so load can find it
        ws.cell(row=r, column=3, value=desc).border = _THIN_BORDER
        ws.cell(row=r, column=3).fill = _fill(_WHITE)

        # hidden key column D (used by loader)
        ws.cell(row=r, column=4, value=key)
        ws.column_dimensions["D"].hidden = True

    # Mode dropdown validation
    dv_mode = DataValidation(
        type="list",
        formula1='"full,partial,max_shift,custom"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_mode.error = "Must be: full, partial, max_shift, or custom"
    dv_mode.errorTitle = "Invalid mode"
    ws.add_data_validation(dv_mode)
    dv_mode.add(ws.cell(row=3, column=2))   # mode row

    # include_irs dropdown — find its row by scanning hidden key column
    dv_bool_cfg = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv_bool_cfg)
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.column == 4 and cell.value == "include_irs":
                dv_bool_cfg.add(ws.cell(row=cell.row, column=2))
                break

    # Column widths
    _set_col_width(ws, "A", 28)
    _set_col_width(ws, "B", 18)
    _set_col_width(ws, "C", 60)

    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Products sheet
# ─────────────────────────────────────────────────────────────────────────────

_PROD_HEADERS = [
    ("product_code", "Product Code",    10),
    ("bs_side",      "Side",            8),
    ("product_name", "Name",            26),
    ("currency",     "CCY",             7),
    ("current_pct",  "Current %",       10),
    ("fixed",        "Fixed",           8),
    ("lb_pct",       "LB %",            9),
    ("ub_pct",       "UB %",            9),
    ("notes",        "Notes",           28),
]


def _write_products_sheet(ws, prod_rows: list[dict]) -> None:
    """Write the 'products' sheet."""
    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(_PROD_HEADERS))}1")
    title = ws["A1"]
    title.value = (
        "Per-Product Bounds  —  fixed=TRUE pins weight at current in any mode; "
        "lb_pct / ub_pct apply in 'custom' mode"
    )
    title.font  = Font(bold=True, size=11, color=_WHITE)
    title.fill  = _fill(_BLUE_HDR)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Column headers
    for col, (_, label, width) in enumerate(_PROD_HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=label)
        cell.font      = _hdr_font()
        cell.fill      = _fill(_GREEN_HDR)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _THIN_BORDER
        _set_col_width(ws, get_column_letter(col), width)

    # Data rows
    dv_bool = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv_bool)

    keys = [k for k, _, _ in _PROD_HEADERS]
    for r, row in enumerate(prod_rows, start=3):
        is_equity = str(row.get("bs_side", "")) == "E"
        row_fill  = _fill(_ORANGE if is_equity else _LT_GREEN)

        for col, key in enumerate(keys, 1):
            cell       = ws.cell(row=r, column=col, value=row.get(key, ""))
            cell.border = _THIN_BORDER
            cell.fill   = row_fill

            # Formatting
            if key in ("current_pct", "lb_pct", "ub_pct"):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")
            elif key == "fixed":
                cell.alignment = Alignment(horizontal="center")
                dv_bool.add(cell)
            elif key in ("product_code", "bs_side", "currency"):
                cell.alignment = Alignment(horizontal="center")

        # Protect reference columns (not enforced, just visual)
        for col in (1, 2, 3, 4, 5):   # read-only reference columns
            ws.cell(row=r, column=col).font = Font(color="595959", size=10)

    # Legend note below the table
    note_row = len(prod_rows) + 4
    ws.cell(row=note_row, column=1, value=(
        "LB % / UB % = lower / upper weight bound as % of total_assets  "
        "(e.g. 5.00 means min 5% of total assets).  "
        "Only used in 'custom' mode.  "
        "Equity rows are always pinned regardless of fixed flag."
    ))
    ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="595959")
    ws.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row, end_column=len(_PROD_HEADERS)
    )

    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Public: create template
# ─────────────────────────────────────────────────────────────────────────────

def create_optimizer_excel(
    params: BalanceSheetParams,
    output_path: str,
    defaults: dict | None = None,
) -> None:
    """Generate a formatted Excel optimizer configuration template.

    Parameters
    ----------
    params      : BalanceSheetParams — current balance sheet (provides product list)
    output_path : path to write the .xlsx file
    defaults    : optional dict overriding default config values
                  (keys match _CONFIG_ROWS keys: 'mode', 'tier1_capital', etc.)
    """
    defaults = defaults or {}

    # ── Build product list ────────────────────────────────────────────────────
    ta = params.total_assets
    seen: dict[tuple[str, str], dict] = {}
    for pc, side, name, ccy, bal in zip(
        params.product_code, params.bs_side, params.product_name,
        params.currency, params.balance_arr,
    ):
        key = (str(pc), str(side))
        if key not in seen:
            seen[key] = {
                "product_code": str(pc),
                "bs_side":      str(side),
                "product_name": str(name),
                "currency":     str(ccy),
                "current_pct":  0.0,
            }
        seen[key]["current_pct"] += float(bal) / ta * 100.0

    # Fill default bounds: equity pinned, others free
    prod_rows = []
    for key, row in seen.items():
        _, side = key
        is_equity = side == "E"
        cpct = round(row["current_pct"], 4)
        row["fixed"]  = True if is_equity else False
        row["lb_pct"] = round(cpct, 4) if is_equity else 0.0
        row["ub_pct"] = round(cpct, 4) if is_equity else 100.0
        row["notes"]  = "equity — always pinned" if is_equity else ""
        prod_rows.append(row)

    # ── Write workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    ws_cfg  = wb.create_sheet("config")
    ws_prod = wb.create_sheet("products")

    _write_config_sheet(ws_cfg,  defaults)
    _write_products_sheet(ws_prod, prod_rows)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"Optimizer config written to: {output_path}")
    print(f"  {len(prod_rows)} products  |  edit 'config' sheet for global params")
    print(f"  Modes: full / partial / max_shift / custom")
    print(f"  In 'custom' mode: set lb_pct / ub_pct per product for asymmetric bounds")


# ─────────────────────────────────────────────────────────────────────────────
# Public: load config
# ─────────────────────────────────────────────────────────────────────────────

def load_optimizer_config(
    excel_path: str,
    params: BalanceSheetParams,
) -> OptimizationConfig:
    """Parse an optimizer Excel config into an OptimizationConfig.

    Parameters
    ----------
    excel_path : path to the .xlsx file (created by create_optimizer_excel)
    params     : BalanceSheetParams — used to resolve total_assets for bound conversion

    Returns
    -------
    OptimizationConfig ready to pass to optimize_nii()
    """
    # ── Global config ─────────────────────────────────────────────────────────
    cfg_df = pd.read_excel(
        excel_path, sheet_name="config",
        header=None, usecols=[1, 3],   # col B (value) + col D (key)
    )
    cfg_df.columns = ["value", "key"]
    cfg_df = cfg_df.dropna(subset=["key"])
    cfg_map = dict(zip(cfg_df["key"].astype(str), cfg_df["value"]))

    def _get(key: str, default, cast=None):
        raw = cfg_map.get(key, default)
        return cast(raw) if (cast is not None and raw is not None) else raw

    mode           = str(_get("mode", "max_shift"))
    tier1_capital  = float(_get("tier1_capital", 0.0))      # 0 = auto from equity
    sot_eve_floor  = float(_get("sot_eve_floor", -15.0))
    sot_nii_floor  = float(_get("sot_nii_floor", -5.0))
    sot_eve_buffer = float(_get("sot_eve_buffer", 3.0))
    sot_nii_buffer = float(_get("sot_nii_buffer", 0.0))
    _irs_raw       = _get("include_irs", True)
    include_irs    = (str(_irs_raw).strip().upper() != "FALSE") if _irs_raw is not None else True
    min_lcr        = float(_get("min_lcr", 1.0))
    min_nsfr       = float(_get("min_nsfr", 1.0))
    min_t1_rwa     = float(_get("min_t1_rwa", 0.0))
    max_shift_pct  = float(_get("max_shift_pct", 3.0))
    max_iter       = int(_get("max_iter", 500))

    # ── Per-product bounds ────────────────────────────────────────────────────
    prod_df = pd.read_excel(excel_path, sheet_name="products", header=1)
    # header=1 reads row 2 as headers (row 1 is the title)

    # Normalise column names to lower-case
    prod_df.columns = [str(c).strip().lower().replace(" ", "_") for c in prod_df.columns]
    # Map friendly display names (from Excel header row 2) back to code keys.
    # After .lower().replace(" ", "_"), headers become e.g. "current_%" or "lb_%".
    _col_remap = {
        "product_code": "product_code",
        "side":         "bs_side",
        "name":         "product_name",
        "ccy":          "currency",
        "current_%":    "current_pct",
        "fixed":        "fixed",
        "lb_%":         "lb_pct",
        "ub_%":         "ub_pct",
        "notes":        "notes",
    }
    prod_df.rename(columns=_col_remap, inplace=True)
    # Fallback: if the header was "LB %" the normalised form is "lb_%" which maps correctly.
    # If still missing after rename, alias from original column names.
    if "lb_pct" not in prod_df.columns:
        for raw in prod_df.columns:
            if "lb" in raw:
                prod_df.rename(columns={raw: "lb_pct"}, inplace=True)
                break
    if "ub_pct" not in prod_df.columns:
        for raw in prod_df.columns:
            if "ub" in raw:
                prod_df.rename(columns={raw: "ub_pct"}, inplace=True)
                break

    # Keep only valid product rows (bs_side must be A, L, or E)
    prod_df["product_code"] = prod_df["product_code"].astype(str).str.strip()
    prod_df["bs_side"]      = prod_df["bs_side"].astype(str).str.strip()
    prod_df = prod_df[prod_df["bs_side"].isin(["A", "L", "E"])].copy()

    fixed_products: set[tuple[str, str]] = set()
    custom_bounds:  dict[tuple[str, str], tuple[float, float]] = {}

    for _, row in prod_df.iterrows():
        pc   = str(row["product_code"]).strip()
        side = str(row["bs_side"]).strip()
        key  = (pc, side)

        # Fixed flag: accept TRUE/FALSE string or bool/1/0
        fixed_raw = row.get("fixed", False)
        if isinstance(fixed_raw, str):
            is_fixed = fixed_raw.strip().upper() == "TRUE"
        else:
            is_fixed = bool(fixed_raw)

        if is_fixed:
            fixed_products.add(key)

        # Custom bounds (used in 'custom' mode; loaded for all modes)
        try:
            lb = float(row.get("lb_pct", 0.0)) / 100.0
            ub = float(row.get("ub_pct", 100.0)) / 100.0
        except (TypeError, ValueError):
            lb, ub = 0.0, 1.0
        custom_bounds[key] = (max(0.0, lb), min(1.0, ub))

    return OptimizationConfig(
        mode           = mode,
        tier1_capital  = tier1_capital,
        fixed_products = fixed_products,
        custom_bounds  = custom_bounds,
        max_shift      = max_shift_pct / 100.0,
        sot_eve_floor  = sot_eve_floor,
        sot_nii_floor  = sot_nii_floor,
        sot_eve_buffer = sot_eve_buffer,
        sot_nii_buffer = sot_nii_buffer,
        include_irs    = include_irs,
        min_lcr        = min_lcr,
        min_nsfr       = min_nsfr,
        min_t1_rwa     = min_t1_rwa,
        max_iter       = max_iter,
    )
