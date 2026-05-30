"""nii_formula_example.py
=======================
Builds an Excel workbook illustrating the proposed formula-based NII
calculation for one fixed cohort and one floating cohort.

Output: optimize_prep/output/nii_formula_example.xlsx
"""
from __future__ import annotations

import os
import numpy as np
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ─── Load data ───────────────────────────────────────────────────────────────
HERE    = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "..", "output")

params  = np.load(os.path.join(OUT_DIR, "product_params.npz"), allow_pickle=True)
curves  = np.load(os.path.join(OUT_DIR, "curve_tensors.npz"),  allow_pickle=True)

codes     = params["product_code"].astype(str)
cohort    = params["is_cohort"].astype(bool)
rt        = params["rate_type"].astype(str)
coupon_a  = params["coupon_rate"].astype(float)
rep_m_a   = params["repricing_tenor_m"].astype(float)
nii_unit  = params["nii_unit_rate"].astype(float)
delta_nii = params["delta_nii_unit"].astype(float)  # (n, S)
balance   = params["balance_arr"].astype(float)
sign_a    = params["sign"].astype(float)
floor_a   = params["client_floor"].astype(float)
cap_a     = params["client_cap"].astype(float)
coeff_a_a = params["coeff_a"].astype(float)
cohort_id = params["cohort_id"].astype(str)
scen_ids  = list(params["scenario_ids"].astype(str))

fwd_rates  = curves["fwd_rates"].astype(float)   # (S_c, C, 360)  — curves scenarios
curve_scen = list(curves["scenario_ids"].astype(str))
curve_ccy  = list(curves["currencies"].astype(str))
PLN_IDX    = curve_ccy.index("PLN")

def get_fwd(scenario: str) -> np.ndarray:
    si = curve_scen.index(scenario)
    return fwd_rates[si, PLN_IDX, :]   # (360,)

# ─── Pick example products ────────────────────────────────────────────────────
# Fixed: bond 3000_A_PLN_2020_09 — rep_m=9, coupon 0.3%
i_fix = int(np.where(cohort_id == "3000_A_PLN_2020_09")[0][0])
# Floating: mortgage 1100_A_PLN_2024_12 — rep_m=3
i_flt = int(np.where(cohort_id == "1100_A_PLN_2024_12")[0][0])

HORIZON  = 12          # NII horizon in months
SCENARIOS = ["base", "par_up", "par_dn", "sr_up", "sr_dn"]

# ─── Core formula ─────────────────────────────────────────────────────────────

def avg_fwd(fwd: np.ndarray, m_start: int, m_end: int) -> float:
    """Average of fwd[m_start:m_end] (0-based month indices).
    Returns 0.0 if the window is empty.
    """
    if m_end <= m_start:
        return 0.0
    return float(fwd[m_start:m_end].mean())


def formula_fixed(balance: float, coupon: float, rep_m: float,
                  fwd_shocked: np.ndarray, fwd_base: np.ndarray,
                  coeff_a: float, floor_: float, cap_: float,
                  sign: float) -> dict:
    """
    Fixed cohort NII formula (shocked scenario).

    Pre-repricing  : earns contractual coupon_rate  (rate locked, no rate risk)
    Post-repricing : earns market rate from shocked fwd curve (renewal)

    Returns dict with all intermediate steps.
    """
    pre_m  = min(int(rep_m), HORIZON)            # months locked at coupon
    post_m = max(HORIZON - int(rep_m), 0)        # months at renewed market rate

    nii_pre  = balance * coupon * (pre_m  / 12) * sign

    post_fwd_base    = avg_fwd(fwd_base,    pre_m, HORIZON)
    post_fwd_shocked = avg_fwd(fwd_shocked, pre_m, HORIZON)

    fl = floor_ if np.isfinite(floor_) else -np.inf
    cp = cap_   if np.isfinite(cap_)   else  np.inf
    renewal_rate_base    = float(np.clip(coeff_a * post_fwd_base    , fl, cp))
    renewal_rate_shocked = float(np.clip(coeff_a * post_fwd_shocked , fl, cp))

    nii_post_base    = balance * renewal_rate_base    * (post_m / 12) * sign
    nii_post_shocked = balance * renewal_rate_shocked * (post_m / 12) * sign

    nii_total         = nii_pre + nii_post_shocked
    nii_total_base    = nii_pre + nii_post_base
    delta_nii_formula = nii_post_shocked - nii_post_base

    return dict(
        pre_m=pre_m, post_m=post_m,
        nii_pre=nii_pre,
        post_fwd_base=post_fwd_base, post_fwd_shocked=post_fwd_shocked,
        renewal_rate_base=renewal_rate_base, renewal_rate_shocked=renewal_rate_shocked,
        nii_post_base=nii_post_base, nii_post_shocked=nii_post_shocked,
        nii_base=nii_total_base, nii_shocked=nii_total,
        delta_nii=delta_nii_formula,
    )


def formula_float(balance: float, rep_m: float,
                  fwd_shocked: np.ndarray, fwd_base: np.ndarray,
                  coeff_a: float, floor_: float, cap_: float,
                  sign: float) -> dict:
    """
    Floating cohort NII formula (shocked scenario).

    Pre-repricing  : current rate locked (approximated from base fwd avg over pre period)
    Post-repricing : earns shocked market rate

    delta_NII = change in post-repricing component only.
    """
    pre_m  = min(int(rep_m), HORIZON)
    post_m = max(HORIZON - int(rep_m), 0)

    fl = floor_ if np.isfinite(floor_) else -np.inf
    cp = cap_   if np.isfinite(cap_)   else  np.inf

    pre_fwd_base   = avg_fwd(fwd_base, 0, pre_m)
    pre_rate_base  = float(np.clip(coeff_a * pre_fwd_base, fl, cp))
    nii_pre        = balance * pre_rate_base * (pre_m / 12) * sign

    post_fwd_base    = avg_fwd(fwd_base,    pre_m, HORIZON)
    post_fwd_shocked = avg_fwd(fwd_shocked, pre_m, HORIZON)

    post_rate_base    = float(np.clip(coeff_a * post_fwd_base,    fl, cp))
    post_rate_shocked = float(np.clip(coeff_a * post_fwd_shocked, fl, cp))

    nii_post_base    = balance * post_rate_base    * (post_m / 12) * sign
    nii_post_shocked = balance * post_rate_shocked * (post_m / 12) * sign

    nii_base    = nii_pre + nii_post_base
    nii_shocked = nii_pre + nii_post_shocked
    delta        = nii_post_shocked - nii_post_base

    return dict(
        pre_m=pre_m, post_m=post_m,
        pre_fwd_base=pre_fwd_base, pre_rate_base=pre_rate_base,
        nii_pre=nii_pre,
        post_fwd_base=post_fwd_base, post_fwd_shocked=post_fwd_shocked,
        post_rate_base=post_rate_base, post_rate_shocked=post_rate_shocked,
        nii_post_base=nii_post_base, nii_post_shocked=nii_post_shocked,
        nii_base=nii_base, nii_shocked=nii_shocked,
        delta_nii=delta,
    )


# ─── Pre-compute all scenarios ────────────────────────────────────────────────
fwd_base = get_fwd("base")

results = {}
for sc in SCENARIOS:
    fwd_sc = get_fwd(sc)
    results[sc] = {
        "fix": formula_fixed(
            balance=balance[i_fix],
            coupon=coupon_a[i_fix],
            rep_m=rep_m_a[i_fix],
            fwd_shocked=fwd_sc,
            fwd_base=fwd_base,
            coeff_a=coeff_a_a[i_fix],
            floor_=float(floor_a[i_fix]),
            cap_=float(cap_a[i_fix]),
            sign=float(sign_a[i_fix]),
        ),
        "flt": formula_float(
            balance=balance[i_flt],
            rep_m=rep_m_a[i_flt],
            fwd_shocked=fwd_sc,
            fwd_base=fwd_base,
            coeff_a=coeff_a_a[i_flt],
            floor_=float(floor_a[i_flt]),
            cap_=float(cap_a[i_flt]),
            sign=float(sign_a[i_flt]),
        ),
    }

# Level 1 comparison
base_sc_idx = scen_ids.index  # callable
def level1_delta(i: int, scenario: str) -> float:
    if scenario == "base":
        return float(balance[i] * nii_unit[i])
    s = scen_ids.index(scenario)
    return float(balance[i] * delta_nii[i, s])

# ─── Excel builder ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "NII Formula Example"

# ── Style helpers ──────────────────────────────────────────────────────────────
HDR_DARK   = PatternFill("solid", fgColor="1F3864")
HDR_BLUE   = PatternFill("solid", fgColor="2E75B6")
HDR_GREEN  = PatternFill("solid", fgColor="375623")
HDR_ORANGE = PatternFill("solid", fgColor="833C00")
HDR_GREY   = PatternFill("solid", fgColor="595959")
ROW_LIGHT  = PatternFill("solid", fgColor="EBF3FB")
ROW_GREEN  = PatternFill("solid", fgColor="E2EFDA")
ROW_ORANGE = PatternFill("solid", fgColor="FCE4D6")
WHITE      = PatternFill("solid", fgColor="FFFFFF")

WHITE_FT  = Font(color="FFFFFF", bold=True, size=10)
DARK_FT   = Font(bold=True, size=10)
NORM_FT   = Font(size=10)
ITALIC_FT = Font(italic=True, size=9, color="595959")

thin = Side(style="thin", color="BDC3C7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

PCT6  = "0.0000%"
PCT2  = "0.00%"
NUM0  = "#,##0"
NUM2  = "#,##0.00"

def cell(ws, r, c, val=None, *, fill=None, font=None, align="left",
         fmt=None, bold=False, border=True):
    cl = ws.cell(row=r, column=c, value=val)
    if fill:  cl.fill   = fill
    if font:  cl.font   = font
    elif bold: cl.font  = Font(bold=True, size=10)
    else:      cl.font  = NORM_FT
    cl.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if fmt:   cl.number_format = fmt
    if border: cl.border = BORDER
    return cl


def hdr(ws, r, c, val, fill, span=1):
    cl = cell(ws, r, c, val, fill=fill, font=WHITE_FT, align="center")
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r, end_column=c + span - 1)
    return cl


def merge_label(ws, r, c, val, fill=None, font=None, span_r=1, span_c=1):
    cl = ws.cell(row=r, column=c, value=val)
    cl.fill      = fill or WHITE
    cl.font      = font or NORM_FT
    cl.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    cl.border    = BORDER
    if span_r > 1 or span_c > 1:
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r + span_r - 1, end_column=c + span_c - 1)
    return cl


# ─── Section 1: Product parameters ────────────────────────────────────────────
r = 1
ws.row_dimensions[r].height = 22
merge_label(ws, r, 1, "NII FORMULA — WORKED EXAMPLE", fill=HDR_DARK,
            font=Font(color="FFFFFF", bold=True, size=13), span_c=20)
r += 1

ws.row_dimensions[r].height = 8   # spacer

r += 1
ws.row_dimensions[r].height = 18
hdr(ws, r, 1, "PRODUCT PARAMETERS", HDR_GREY, span=8)
r += 1

param_hdr_fill = PatternFill("solid", fgColor="D9D9D9")
for c, lbl in enumerate(["Parameter", "Value (Fixed)", "Value (Float)"], start=1):
    cell(ws, r, c, lbl, fill=param_hdr_fill, bold=True, align="center")

rows_params = [
    ("cohort_id",         cohort_id[i_fix],        cohort_id[i_flt]),
    ("product",           "bond_fixed (3000)",      "mortgage_float (1100)"),
    ("bs_side",           "Asset",                  "Asset"),
    ("balance (PLN)",     balance[i_fix],           balance[i_flt]),
    ("rate_type",         "Fixed (F)",              "Floating (V)"),
    ("coupon_rate",       coupon_a[i_fix],          "n/a (no contractual rate)"),
    ("repricing_tenor_m", rep_m_a[i_fix],           rep_m_a[i_flt]),
    ("coeff_a",           coeff_a_a[i_fix],         coeff_a_a[i_flt]),
    ("floor",             "none",                   "none"),
    ("nii_unit_rate (L1)",nii_unit[i_fix],          nii_unit[i_flt]),
]

r += 1
fills_p = [WHITE, ROW_LIGHT]
for pi, (pname, pfix, pflt) in enumerate(rows_params):
    fl = fills_p[pi % 2]
    cell(ws, r, 1, pname,  fill=fl, bold=True)
    fmt_fix = PCT6 if isinstance(pfix, float) and abs(pfix) < 2 else (NUM0 if isinstance(pfix, float) and abs(pfix) > 1000 else NUM2)
    fmt_flt = PCT6 if isinstance(pflt, float) and abs(pflt) < 2 else (NUM0 if isinstance(pflt, float) and abs(pflt) > 1000 else NUM2)
    cell(ws, r, 2, pfix,   fill=fl, fmt=fmt_fix if isinstance(pfix, float) else None)
    cell(ws, r, 3, pflt,   fill=fl, fmt=fmt_flt if isinstance(pflt, float) else None)
    r += 1

# ─── Section 2: Forward curve ─────────────────────────────────────────────────
r += 1
ws.row_dimensions[r].height = 18
hdr(ws, r, 1, "FORWARD CURVE — PLN (annualised, monthly)",
    HDR_GREY, span=len(SCENARIOS) + 1)
r += 1

cell(ws, r, 1, "Month", fill=param_hdr_fill, bold=True, align="center")
for ci, sc in enumerate(SCENARIOS, start=2):
    cell(ws, r, ci, sc, fill=param_hdr_fill, bold=True, align="center")
r += 1

for m in range(HORIZON):
    fl = fills_p[m % 2]
    cell(ws, r, 1, m + 1, fill=fl, align="center")
    for ci, sc in enumerate(SCENARIOS, start=2):
        fv = float(get_fwd(sc)[m])
        cell(ws, r, ci, fv, fill=fl, fmt=PCT6)
    r += 1

# ─── Section 3: Step-by-step fixed cohort ─────────────────────────────────────
r += 1
ws.row_dimensions[r].height = 18
hdr(ws, r, 1,
    f"FIXED COHORT  ·  {cohort_id[i_fix]}  ·  coupon {coupon_a[i_fix]:.4%}  ·  "
    f"rep_m {rep_m_a[i_fix]:.0f}m  ·  balance {balance[i_fix]:,.0f} PLN",
    HDR_GREEN, span=len(SCENARIOS) + 2)
r += 1

fix_row_defs = [
    ("pre_months",           "min(rep_m, 12)",                       "pre_m",             None),
    ("post_months",          "max(12 - rep_m, 0)",                   "post_m",            None),
    ("NII pre-repricing",    "balance × coupon × pre_m/12",          "nii_pre",           NUM0),
    ("avg fwd BASE [pre..12]","avg of base fwd[pre_m : 12]",          "post_fwd_base",     PCT6),
    ("avg fwd SHOCKED [pre..12]","avg of scenario fwd[pre_m : 12]",  "post_fwd_shocked",  PCT6),
    ("renewal rate BASE",    "coeff_a × avg_fwd_base",               "renewal_rate_base", PCT6),
    ("renewal rate SHOCKED", "coeff_a × avg_fwd_shocked",            "renewal_rate_shocked", PCT6),
    ("NII post BASE",        "balance × renewal_base × post_m/12",   "nii_post_base",     NUM0),
    ("NII post SHOCKED",     "balance × renewal_shocked × post_m/12","nii_post_shocked",  NUM0),
    ("── TOTAL NII (formula)──","= NII_pre + NII_post_shocked",       "nii_shocked",       NUM0),
    ("── delta NII (formula)──","= NII_post_shocked - NII_post_base", "delta_nii",         NUM0),
    ("── Level 1 NII ────────","balance × nii_unit_rate  [L1 base]",  "_l1_nii",           NUM0),
    ("── Level 1 delta NII ──","balance × delta_nii_unit  [L1 delta]","_l1_delta",         NUM0),
]

cell(ws, r, 1, "Step", fill=param_hdr_fill, bold=True)
cell(ws, r, 2, "Formula", fill=param_hdr_fill, bold=True)
for ci, sc in enumerate(SCENARIOS, start=3):
    cell(ws, r, ci, sc, fill=param_hdr_fill, bold=True, align="center")
r += 1

for ri, (step, formula, key, fmt) in enumerate(fix_row_defs):
    fl = fills_p[ri % 2]
    highlight = step.startswith("──")
    fill_use  = ROW_GREEN if highlight else fl
    font_use  = Font(bold=True, size=10) if highlight else NORM_FT
    cell(ws, r, 1, step,    fill=fill_use, font=font_use)
    cell(ws, r, 2, formula, fill=fill_use, font=ITALIC_FT if not highlight else font_use)
    for ci, sc in enumerate(SCENARIOS, start=3):
        d = results[sc]["fix"]
        if key == "_l1_nii":
            v = float(balance[i_fix] * nii_unit[i_fix])
        elif key == "_l1_delta":
            v = 0.0 if sc == "base" else float(balance[i_fix] * delta_nii[i_fix, scen_ids.index(sc)])
        else:
            v = d[key]
        cell(ws, r, ci, v, fill=fill_use, font=font_use, fmt=fmt)
    r += 1

# ─── Section 4: Step-by-step floating cohort ─────────────────────────────────
r += 1
ws.row_dimensions[r].height = 18
hdr(ws, r, 1,
    f"FLOATING COHORT  ·  {cohort_id[i_flt]}  ·  rep_m {rep_m_a[i_flt]:.0f}m  ·  "
    f"balance {balance[i_flt]:,.0f} PLN",
    HDR_ORANGE, span=len(SCENARIOS) + 2)
r += 1

flt_row_defs = [
    ("pre_months",            "min(rep_m, 12)",                          "pre_m",             None),
    ("post_months",           "max(12 - rep_m, 0)",                      "post_m",            None),
    ("avg fwd BASE [0..pre]", "avg of base fwd[0 : pre_m]",              "pre_fwd_base",      PCT6),
    ("pre rate (locked)",     "coeff_a × avg_pre_fwd_base",              "pre_rate_base",     PCT6),
    ("NII pre-repricing",     "balance × pre_rate × pre_m/12",           "nii_pre",           NUM0),
    ("avg fwd BASE [pre..12]","avg of base fwd[pre_m : 12]",             "post_fwd_base",     PCT6),
    ("avg fwd SHOCKED [pre..12]","avg of scenario fwd[pre_m : 12]",      "post_fwd_shocked",  PCT6),
    ("post rate BASE",        "coeff_a × avg_post_fwd_base",             "post_rate_base",    PCT6),
    ("post rate SHOCKED",     "coeff_a × avg_post_fwd_shocked",          "post_rate_shocked", PCT6),
    ("NII post BASE",         "balance × post_rate_base × post_m/12",    "nii_post_base",     NUM0),
    ("NII post SHOCKED",      "balance × post_rate_shocked × post_m/12", "nii_post_shocked",  NUM0),
    ("── TOTAL NII (formula)──","= NII_pre + NII_post_shocked",           "nii_shocked",       NUM0),
    ("── delta NII (formula)──","= NII_post_shocked - NII_post_base",     "delta_nii",         NUM0),
    ("── Level 1 NII ────────", "balance × nii_unit_rate  [L1 base]",    "_l1_nii",           NUM0),
    ("── Level 1 delta NII ──", "balance × delta_nii_unit  [L1 delta]",  "_l1_delta",         NUM0),
]

cell(ws, r, 1, "Step", fill=param_hdr_fill, bold=True)
cell(ws, r, 2, "Formula", fill=param_hdr_fill, bold=True)
for ci, sc in enumerate(SCENARIOS, start=3):
    cell(ws, r, ci, sc, fill=param_hdr_fill, bold=True, align="center")
r += 1

for ri, (step, formula, key, fmt) in enumerate(flt_row_defs):
    fl = fills_p[ri % 2]
    highlight = step.startswith("──")
    fill_use  = ROW_ORANGE if highlight else fl
    font_use  = Font(bold=True, size=10) if highlight else NORM_FT
    cell(ws, r, 1, step,    fill=fill_use, font=font_use)
    cell(ws, r, 2, formula, fill=fill_use, font=ITALIC_FT if not highlight else font_use)
    for ci, sc in enumerate(SCENARIOS, start=3):
        d = results[sc]["flt"]
        if key == "_l1_nii":
            v = float(balance[i_flt] * nii_unit[i_flt])
        elif key == "_l1_delta":
            v = 0.0 if sc == "base" else float(balance[i_flt] * delta_nii[i_flt, scen_ids.index(sc)])
        else:
            v = d[key]
        cell(ws, r, ci, v, fill=fill_use, font=font_use, fmt=fmt)
    r += 1

# ─── Section 5: Summary comparison ───────────────────────────────────────────
r += 1
ws.row_dimensions[r].height = 18
hdr(ws, r, 1, "SUMMARY — Formula vs Level 1 delta_NII", HDR_BLUE,
    span=len(SCENARIOS) + 2)
r += 1
cell(ws, r, 1, "Product", fill=param_hdr_fill, bold=True)
cell(ws, r, 2, "Method",  fill=param_hdr_fill, bold=True)
for ci, sc in enumerate(SCENARIOS[1:], start=3):   # skip base
    cell(ws, r, ci, sc, fill=param_hdr_fill, bold=True, align="center")
r += 1

for label, idx, ptype in [
    ("bond_fixed 3000", i_fix, "fix"),
    ("mortgage_float 1100", i_flt, "flt"),
]:
    for method in ("Formula", "Level 1"):
        fl = ROW_GREEN if ptype == "fix" else ROW_ORANGE
        cell(ws, r, 1, label,  fill=fl, bold=(method == "Formula"))
        cell(ws, r, 2, method, fill=fl, bold=(method == "Formula"))
        for ci, sc in enumerate(SCENARIOS[1:], start=3):
            if method == "Formula":
                v = results[sc][ptype]["delta_nii"]
            else:
                v = float(balance[idx] * delta_nii[idx, scen_ids.index(sc)])
            cell(ws, r, ci, v, fill=fl,
                 bold=(method == "Formula"), fmt=NUM0)
        r += 1

# ─── Column widths ────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 42
for c in range(3, 3 + len(SCENARIOS)):
    ws.column_dimensions[get_column_letter(c)].width = 18

ws.freeze_panes = "C1"

out_path = os.path.join(OUT_DIR, "nii_formula_example.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")

# ─── Print key numbers to terminal ────────────────────────────────────────────
print("\n=== FIXED COHORT: bond_fixed 3000_A_PLN_2020_09 ===")
print(f"  coupon={coupon_a[i_fix]:.4%}  rep_m={rep_m_a[i_fix]:.0f}m  balance={balance[i_fix]:,.0f}")
for sc in SCENARIOS:
    d = results[sc]["fix"]
    l1 = 0.0 if sc == "base" else float(balance[i_fix] * delta_nii[i_fix, scen_ids.index(sc)])
    tag = f"NII={d['nii_shocked']:,.0f}" if sc == "base" else f"delta={d['delta_nii']:,.0f}  L1={l1:,.0f}"
    print(f"  {sc:<8}: {tag}")

print("\n=== FLOATING COHORT: mortgage_float 1100_A_PLN_2024_12 ===")
print(f"  rep_m={rep_m_a[i_flt]:.0f}m  balance={balance[i_flt]:,.0f}")
for sc in SCENARIOS:
    d = results[sc]["flt"]
    l1 = 0.0 if sc == "base" else float(balance[i_flt] * delta_nii[i_flt, scen_ids.index(sc)])
    tag = f"NII={d['nii_shocked']:,.0f}" if sc == "base" else f"delta={d['delta_nii']:,.0f}  L1={l1:,.0f}"
    print(f"  {sc:<8}: {tag}")
